"""Excel file processor for extracting specific sheets."""
import os
import pandas as pd
from openpyxl import load_workbook
from typing import Optional, List
from config import Config


class ExcelProcessor:
    """Processor for extracting and analyzing Excel files."""
    
    def __init__(self, target_sheet: str = None):
        """
        Initialize Excel processor.
        
        Args:
            target_sheet: Name of the sheet to extract (default: from Config)
        """
        self.target_sheet = target_sheet or Config.TARGET_SHEET_NAME
    
    def extract_sheet(self, filepath: str) -> Optional[pd.DataFrame]:
        """
        Extract a specific sheet from an Excel file.
        
        Args:
            filepath: Path to the Excel file
        
        Returns:
            DataFrame containing the sheet data, or None if sheet not found
        """
        if not os.path.exists(filepath):
            print(f"Error: File not found: {filepath}")
            return None
        
        try:
            # Load workbook to check available sheets
            workbook = load_workbook(filepath, read_only=True)
            available_sheets = workbook.sheetnames
            
            print(f"\nProcessing: {os.path.basename(filepath)}")
            print(f"Available sheets: {', '.join(available_sheets)}")
            
            if self.target_sheet not in available_sheets:
                print(f"⚠ Warning: Sheet '{self.target_sheet}' not found in {filepath}")
                print(f"Available sheets are: {', '.join(available_sheets)}")
                return None
            
            # Read the target sheet
            df = pd.read_excel(filepath, sheet_name=self.target_sheet)
            
            print(f"✓ Successfully extracted sheet '{self.target_sheet}'")
            print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")
            
            return df
        
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            return None
    
    def get_sheet_info(self, filepath: str) -> dict:
        """
        Get information about all sheets in an Excel file.
        
        Args:
            filepath: Path to the Excel file
        
        Returns:
            Dictionary with sheet names and row counts
        """
        try:
            workbook = load_workbook(filepath, read_only=True)
            sheet_info = {}
            
            for sheet_name in workbook.sheetnames:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                sheet_info[sheet_name] = {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns)
                }
            
            return sheet_info
        
        except Exception as e:
            print(f"Error getting sheet info: {e}")
            return {}
    
    def display_preview(self, df: pd.DataFrame, rows: int = 5):
        """
        Display a preview of the DataFrame.
        
        Args:
            df: DataFrame to preview
            rows: Number of rows to display
        """
        if df is None or df.empty:
            print("No data to display")
            return
        
        print(f"\n{'='*80}")
        print(f"Data Preview (first {rows} rows):")
        print(f"{'='*80}")
        print(df.head(rows).to_string())
        print(f"{'='*80}")
        print(f"\nColumn Names: {', '.join(df.columns)}")
        print(f"Total Rows: {len(df)}")
        print(f"Total Columns: {len(df.columns)}")
    
    def save_to_csv(self, df: pd.DataFrame, output_path: str) -> bool:
        """
        Save DataFrame to CSV file.
        
        Args:
            df: DataFrame to save
            output_path: Path to save CSV file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            df.to_csv(output_path, index=False)
            print(f"✓ Saved to CSV: {output_path}")
            return True
        except Exception as e:
            print(f"Error saving to CSV: {e}")
            return False
    
    def get_summary_stats(self, df: pd.DataFrame) -> dict:
        """
        Get summary statistics for the DataFrame.
        
        Args:
            df: DataFrame to analyze
        
        Returns:
            Dictionary with summary statistics
        """
        if df is None or df.empty:
            return {}
        
        return {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'column_names': list(df.columns),
            'numeric_columns': list(df.select_dtypes(include=['number']).columns),
            'text_columns': list(df.select_dtypes(include=['object']).columns),
            'missing_values': df.isnull().sum().to_dict(),
            'memory_usage': df.memory_usage(deep=True).sum()
        }
